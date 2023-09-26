import os
import openpyxl as xl

def printSheet1(sheet)-> None :
	row = sheet.max_row
	column = sheet.max_column

	for i in range(1,row+ 1) :														
		for j in range (1,column+1) :
			cell = sheet.cell(row=i,column=j)
			print(cell.value,end="	")
		print()
	return 
##########################
def printSheet2(sheet)->None :
	cell = sheet['A1':'C6']
	for cell1,cell2,cell3 in cell :
		print(cell1.value," ",cell2.value,cell3.value)
	return 	
##########################
def write() :
	newFile = xl.Workbook()
	sheet = newFile.active
	c1 = sheet.cell(row=1,column=1)
	c1.value = "Hello"																					
	c2 = sheet.cell(row=1,column=2)
	c2.value = "Man !"		
	c3 = sheet['A2']	
	c3.value = "Welcome"																		
	c4 = sheet['B2']	
	c4.value = "Python !"	
	newFile.save(filename="test2.xlsx")
	return 
##########################
def appendData(fileName) :
	file = xl.load_workbook(fileName)
	sheet = file.active
	data = (
	(),
	(None,100,None,60),
	(10,20,30),
	(40,50,60,70)
	)
	for row in data :
		sheet.append(row)
	file.save(fileName)
	
	return
def sum() :
    file = xl.Workbook()
    sheet = file.active
    sheet['A1'] = 200
    sheet['A2'] = 500
    sheet['A3'] = 400
    sheet['A4'] = 300
    sheet['A5'] = 600
    sheet['A6'] = '= SUM(A1:A5)'
    file.save("sum.xlsx")
    return
##########################
def x() :
    
    return 
##########################

##########################
#### 《MAIN FUNCTION》 ###
##########################
path = "/storage/emulated/0/A XL/" 
file = "test.xlsx"
os.chdir(path)

testFile = xl.load_workbook(file)
sheet = testFile.active


